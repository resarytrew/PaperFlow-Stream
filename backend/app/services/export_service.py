"""Export bundles: ZIP, CSV, JSON, XLSX (sections 6.13 and 7.10)."""

from __future__ import annotations

import csv
import io
import json
import logging
import zipfile
from datetime import datetime, timezone
from pathlib import Path

from sqlalchemy import select
from sqlalchemy.orm import Session

from app.models import RecognitionStatus, ScanSession, ScanStatus, ScannedSheet
from app.services.storage import get_storage

logger = logging.getLogger(__name__)


def _answer_verdict(recognition) -> str:
    """RU label of the answer hint for flat exports ('' when not available)."""
    if recognition is None or not recognition.analysis_json:
        return ""
    check = (recognition.analysis_json or {}).get("answerCheck") or {}
    labels = {
        "match": "совпадает",
        "likely": "похоже",
        "mismatch": "отличается",
        "unknown": "",
    }
    return labels.get(check.get("verdict", "unknown"), "")


def _sheet_rows(sheets: list[ScannedSheet]) -> list[dict]:
    rows: list[dict] = []
    for sheet in sheets:
        recognition = sheet.recognition
        review = sheet.review
        rows.append(
            {
                "number": sheet.sequence_number,
                "sheet_uid": sheet.sheet_uid or "",
                "student_external_id": sheet.student.external_id if sheet.student else "",
                "student_name": sheet.student.display_name if sheet.student else "",
                "class": (sheet.student.class_group.name if sheet.student and sheet.student.class_group else ""),
                "task": sheet.task.external_id if sheet.task else "",
                "task_title": sheet.task.title if sheet.task else "",
                "scanned_at": sheet.created_at.isoformat() if sheet.created_at else "",
                "quality": round(sheet.quality_score, 4),
                "sharpness": round(sheet.sharpness_score, 4),
                "glare": round(sheet.glare_score, 4),
                "qr_status": sheet.qr_status,
                "scan_status": sheet.scan_status,
                "duplicate_of": sheet.duplicate_of_id or "",
                "warnings": "; ".join(sheet.warnings or []),
                "recognized_text": (recognition.recognized_text if recognition else ""),
                "ocr_status": (recognition.status if recognition else ""),
                "ocr_confidence": (round(recognition.overall_confidence, 4) if recognition else ""),
                "ocr_provider": (recognition.provider if recognition else ""),
                "ocr_model": (recognition.model_name if recognition else ""),
                "answer_verdict": _answer_verdict(recognition),
                "teacher_text": (review.teacher_text if review else ""),
                "review_decision": (review.decision if review else ""),
                "review_comment": (review.comment if review else ""),
            }
        )
    return rows


def _select_sheets(
    db: Session,
    session_id: int,
    *,
    only_problems: bool = False,
    only_uncertain: bool = False,
    only_corrected: bool = False,
) -> list[ScannedSheet]:
    query = select(ScannedSheet).where(
        ScannedSheet.session_id == session_id,
        ScannedSheet.scan_status != ScanStatus.deleted.value,
    )
    sheets = list(db.execute(query.order_by(ScannedSheet.sequence_number, ScannedSheet.id)).scalars().all())

    if only_problems:
        sheets = [
            s
            for s in sheets
            if s.scan_status
            in (
                ScanStatus.duplicate.value,
                ScanStatus.unidentified.value,
                ScanStatus.low_quality.value,
                ScanStatus.rescan_required.value,
            )
            or s.qr_status != "read"
        ]
    if only_uncertain:
        sheets = [
            s
            for s in sheets
            if s.recognition is not None
            and s.recognition.status
            in (RecognitionStatus.needs_review.value, RecognitionStatus.failed.value)
        ]
    if only_corrected:
        sheets = [s for s in sheets if s.review is not None and s.review.decision == "corrected"]
    return sheets


def export_csv(db: Session, session_id: int, **filters) -> bytes:
    """Flat CSV listing of the works (0.1) including recognised text (0.2)."""
    sheets = _select_sheets(db, session_id, **filters)
    rows = _sheet_rows(sheets)
    buffer = io.StringIO()
    fieldnames = list(rows[0].keys()) if rows else ["number", "sheet_uid", "student_name"]
    writer = csv.DictWriter(buffer, fieldnames=fieldnames, delimiter=";", quoting=csv.QUOTE_MINIMAL)
    writer.writeheader()
    for row in rows:
        writer.writerow(row)
    # BOM so Excel opens Cyrillic correctly
    return b"\xef\xbb\xbf" + buffer.getvalue().encode("utf-8")


def export_json(db: Session, session_id: int, **filters) -> bytes:
    """Technical metadata + per-line recognition (0.2)."""
    session = db.get(ScanSession, session_id)
    sheets = _select_sheets(db, session_id, **filters)

    payload = {
        "exportedAt": datetime.now(timezone.utc).isoformat(),
        "session": {
            "id": session.id if session else session_id,
            "title": session.title if session else "",
            "status": session.status if session else "",
            "class": session.class_group.name if session and session.class_group else None,
            "task": session.task.external_id if session and session.task else None,
            "expectedSheetCount": session.expected_sheet_count if session else 0,
            "startedAt": session.started_at.isoformat() if session and session.started_at else None,
            "completedAt": session.completed_at.isoformat() if session and session.completed_at else None,
        },
        "sheets": [],
    }

    for sheet in sheets:
        recognition = sheet.recognition
        payload["sheets"].append(
            {
                "id": sheet.id,
                "number": sheet.sequence_number,
                "sheetUid": sheet.sheet_uid,
                "studentExternalId": sheet.student.external_id if sheet.student else None,
                "studentName": sheet.student.display_name if sheet.student else None,
                "taskExternalId": sheet.task.external_id if sheet.task else None,
                "scannedAt": sheet.created_at.isoformat() if sheet.created_at else None,
                "qrStatus": sheet.qr_status,
                "qrPayload": sheet.qr_payload,
                "scanStatus": sheet.scan_status,
                "duplicateOfId": sheet.duplicate_of_id,
                "warnings": sheet.warnings,
                "processingTimeMs": sheet.processing_time_ms,
                "metrics": {
                    "quality": round(sheet.quality_score, 4),
                    "sharpness": round(sheet.sharpness_score, 4),
                    "glare": round(sheet.glare_score, 4),
                    "occlusion": round(sheet.occlusion_score, 4),
                    "perspective": round(sheet.perspective_score, 4),
                    "coverage": round(sheet.coverage_score, 4),
                    "motion": round(sheet.motion_score, 4),
                },
                "files": {
                    "source": sheet.source_frame_path,
                    "normalized": sheet.normalized_image_path,
                    "enhanced": sheet.enhanced_image_path,
                    "answerCrop": sheet.answer_crop_path,
                    "thumbnail": sheet.thumbnail_path,
                },
                "recognition": (
                    {
                        "status": recognition.status,
                        "provider": recognition.provider,
                        "modelName": recognition.model_name,
                        "text": recognition.recognized_text,
                        "overallConfidence": round(recognition.overall_confidence, 4),
                        "lines": recognition.line_results_json,
                        "warnings": recognition.warnings,
                        "analysis": recognition.analysis_json,
                        "preprocessVariant": recognition.preprocess_variant,
                        "processingTimeMs": recognition.processing_time_ms,
                        "error": recognition.error_message,
                    }
                    if recognition
                    else None
                ),
                "review": (
                    {
                        "decision": sheet.review.decision,
                        "teacherText": sheet.review.teacher_text,
                        "comment": sheet.review.comment,
                        "reviewedAt": sheet.review.reviewed_at.isoformat(),
                    }
                    if sheet.review
                    else None
                ),
            }
        )
    return json.dumps(payload, ensure_ascii=False, indent=2).encode("utf-8")


def export_xlsx(db: Session, session_id: int, **filters) -> bytes:
    """Spreadsheet with students, tasks, recognised text and confidence (0.2)."""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    sheets = _select_sheets(db, session_id, **filters)
    session = db.get(ScanSession, session_id)

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Работы"

    headers = [
        "№", "sheetId", "ID ученика", "Ученик", "Класс", "Задание",
        "Время", "Качество", "QR", "Статус",
        "Распознанный текст", "Уверенность", "Статус OCR",
        "Текст учителя", "Решение", "Комментарий",
    ]
    worksheet.append(headers)

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="374151")
    for column in range(1, len(headers) + 1):
        cell = worksheet.cell(row=1, column=column)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(vertical="center")

    green = PatternFill("solid", fgColor="DCFCE7")
    yellow = PatternFill("solid", fgColor="FEF9C3")
    red = PatternFill("solid", fgColor="FEE2E2")

    for sheet in sheets:
        recognition = sheet.recognition
        review = sheet.review
        confidence = recognition.overall_confidence if recognition else None
        worksheet.append(
            [
                sheet.sequence_number,
                sheet.sheet_uid or "",
                sheet.student.external_id if sheet.student else "",
                sheet.student.display_name if sheet.student else "",
                (sheet.student.class_group.name if sheet.student and sheet.student.class_group else ""),
                sheet.task.external_id if sheet.task else "",
                sheet.created_at.strftime("%Y-%m-%d %H:%M:%S") if sheet.created_at else "",
                round(sheet.quality_score, 3),
                sheet.qr_status,
                sheet.scan_status,
                recognition.recognized_text if recognition else "",
                round(confidence, 3) if confidence is not None else "",
                recognition.status if recognition else "",
                review.teacher_text if review else "",
                review.decision if review else "",
                review.comment if review else "",
            ]
        )
        row_index = worksheet.max_row
        if confidence is not None:
            fill = green if confidence >= 0.85 else (yellow if confidence >= 0.60 else red)
            worksheet.cell(row=row_index, column=12).fill = fill
        worksheet.cell(row=row_index, column=11).alignment = Alignment(wrap_text=True, vertical="top")

    widths = [5, 24, 12, 24, 8, 18, 19, 9, 12, 14, 46, 11, 14, 40, 14, 26]
    for index, width in enumerate(widths, start=1):
        worksheet.column_dimensions[get_column_letter(index)].width = width
    worksheet.freeze_panes = "A2"

    # Summary sheet
    summary = workbook.create_sheet("Сводка")
    summary.append(["Сессия", session.title if session else ""])
    summary.append(["Класс", session.class_group.name if session and session.class_group else ""])
    summary.append(["Задание", session.task.title if session and session.task else ""])
    summary.append(["Ожидалось работ", session.expected_sheet_count if session else 0])
    summary.append(["Отсканировано", len(sheets)])
    summary.append(["Экспортировано", datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M UTC")])
    summary.append([])
    summary.append(["Примечание", "Уверенность распознавания — техническая метрика, а не оценка работы."])
    summary.column_dimensions["A"].width = 22
    summary.column_dimensions["B"].width = 62

    buffer = io.BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def export_zip(
    db: Session,
    session_id: int,
    *,
    include_source: bool = False,
    include_enhanced: bool = True,
    include_crops: bool = True,
    thumbnails_only: bool = False,
    **filters,
) -> bytes:
    """Bundle images + CSV + JSON into a single archive."""
    sheets = _select_sheets(db, session_id, **filters)
    storage = get_storage()

    buffer = io.BytesIO()
    with zipfile.ZipFile(buffer, "w", zipfile.ZIP_DEFLATED, compresslevel=6) as archive:
        for sheet in sheets:
            student = sheet.student.external_id if sheet.student else "unidentified"
            base = f"{sheet.sequence_number:03d}_{student}_{sheet.sheet_uid or sheet.id}"

            targets: list[tuple[str | None, str]] = []
            if thumbnails_only:
                targets.append((sheet.thumbnail_path, f"thumbnails/{base}.jpg"))
            else:
                targets.append((sheet.normalized_image_path, f"normalized/{base}.jpg"))
                if include_enhanced:
                    targets.append((sheet.enhanced_image_path, f"enhanced/{base}.png"))
                if include_crops:
                    targets.append((sheet.answer_crop_path, f"answers/{base}.jpg"))
                if include_source:
                    targets.append((sheet.source_frame_path, f"originals/{base}.jpg"))

            for relative, arcname in targets:
                if not relative:
                    continue
                try:
                    path: Path = storage.absolute(relative)
                    if path.exists():
                        archive.write(path, arcname)
                except Exception as exc:  # a missing file must not break the export
                    logger.warning("skipping %s in export: %s", relative, exc)

        archive.writestr("works.csv", export_csv(db, session_id, **filters))
        archive.writestr("metadata.json", export_json(db, session_id, **filters))
        readme = (
            "PaperFlow Stream — экспорт сессии\n\n"
            "normalized/ — выровненные цветные листы\n"
            "enhanced/   — контрастные Ч/Б версии\n"
            "answers/    — вырезанные области ответа\n"
            "originals/  — исходные кадры камеры (если включены)\n"
            "works.csv   — таблица работ\n"
            "metadata.json — технические метаданные и построчное распознавание\n\n"
            "Уверенность распознавания — техническая метрика, а не оценка работы ученика.\n"
        )
        archive.writestr("README.txt", readme.encode("utf-8"))

    return buffer.getvalue()
